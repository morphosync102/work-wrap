#!/usr/bin/env python3
"""
MD → CATW Excel 転記スクリプト

MDファイルの「## CATW（案件工数）」テーブルを読み込み、
config.yaml で指定した CATW 提出用 Excel ファイルに転記します。

使い方:
  python scripts/transfer_catw.py 2026 3   # 特定の年月
  python scripts/transfer_catw.py           # 引数なしで今月

Excel 構造（CATW / PSA シート共通）:
  C4=Year, C5=Month
  週ブロック開始行: Week1=7, Week2=19, Week3=31, Week4=43, Week5=55, Week6=67
  各ブロック: ヘッダー行 + 7データ行
  列: C=WBS / D=Description / E=AA Type / F=Memo / G=週合計 / H-N=月〜日
"""

import sys
import re
from datetime import date, timedelta
from pathlib import Path

import yaml
from openpyxl import load_workbook

# Excel構造定数（catw_selenium.py と同じ値）
WEEK_START_ROWS = {1: 7, 2: 19, 3: 31, 4: 43, 5: 55, 6: 67}
DATA_ROW_COUNT = 7

COL_WBS = 3          # C: WBS Element
COL_DESCRIPTION = 4  # D: Description
COL_AA_TYPE = 5      # E: Attend/Absence Type
COL_MEMO = 6         # F: Memo
# G列(7) = 週合計（数式のため書き込まない）
COL_MON = 8          # H: 月曜
COL_TUE = 9          # I: 火曜
COL_WED = 10         # J: 水曜
COL_THU = 11         # K: 木曜
COL_FRI = 12         # L: 金曜
COL_SAT = 13         # M: 土曜
COL_SUN = 14         # N: 日曜

# weekday() (0=Mon..6=Sun) → 列番号
WEEKDAY_TO_COL = {
    0: COL_MON, 1: COL_TUE, 2: COL_WED,
    3: COL_THU, 4: COL_FRI, 5: COL_SAT, 6: COL_SUN,
}


def load_config(path: Path = Path("config.yaml")) -> dict:
    if not path.exists():
        print(f"[エラー] 設定ファイルが見つかりません: {path}")
        print("  cp config.example.yaml config.yaml  を実行して設定してください。")
        sys.exit(1)
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def get_week_num(d: date) -> int:
    """日付から月内週番号(1-6)を計算する。

    週は月曜始まり。1日を含む週を Week 1 とし、
    1日が月曜でなければ直前の月曜を Week 1 月曜とする。
    """
    first_day = date(d.year, d.month, 1)
    week1_monday = first_day - timedelta(days=first_day.weekday())
    week_num = (d - week1_monday).days // 7 + 1
    return max(1, min(6, week_num))


def parse_hours(cell: str) -> float:
    """セル文字列から工数(float)を解析。空文字・'-'は0。"""
    cell = cell.strip()
    if not cell or cell in ("-", " - "):
        return 0.0
    try:
        return float(cell)
    except ValueError:
        return 0.0


def parse_date_label(label: str, year: int, month: int) -> date | None:
    """日付ラベル (M/D ...) を date オブジェクトに変換。**太字**も対応。"""
    label = label.strip().strip("*").strip()
    m = re.match(r"(\d+)/(\d+)", label)
    if not m:
        return None
    try:
        return date(year, int(m.group(1)), int(m.group(2)))
    except ValueError:
        return None


def parse_catw_table(md_text: str, year: int, month: int, projects: list) -> dict:
    """MDの CATW テーブルをパースする。

    Returns:
        {project_name: {date: hours}}
    """
    m = re.search(r"## CATW（案件工数）\n(.*?)(?=\n## |\Z)", md_text, re.DOTALL)
    if not m:
        print("[警告] 'CATW（案件工数）' セクションが見つかりません")
        return {}

    section = m.group(1)
    project_names = [p["name"] for p in projects]
    result: dict[str, dict[date, float]] = {p: {} for p in project_names}

    for line in section.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if not cells or cells[0] in ("日付", "------", ""):
            continue
        if re.match(r"-{3,}", cells[0]):
            continue

        d = parse_date_label(cells[0], year, month)
        if d is None or d.month != month:
            continue

        for i, p_name in enumerate(project_names):
            col_idx = i + 1
            if col_idx < len(cells):
                h = parse_hours(cells[col_idx])
                if h > 0:
                    result[p_name][d] = h

    return result


def build_week_data(project_hours: dict, projects: list) -> dict:
    """project_hours を週ブロック構造に変換する。

    Args:
        project_hours: {project_name: {date: hours}}
        projects: config.yaml の projects リスト

    Returns:
        {week_num: [{wbs, description, aa_type, memo, hours: {weekday: h}}]}
    """
    project_map = {p["name"]: p for p in projects}
    week_data: dict[int, list] = {i: [] for i in range(1, 7)}

    for p_name, date_hours in project_hours.items():
        if not date_hours:
            continue
        p = project_map.get(p_name)
        if not p:
            continue

        # 週ごとにグループ化
        by_week: dict[int, dict[int, float]] = {}
        for d, h in date_hours.items():
            wn = get_week_num(d)
            by_week.setdefault(wn, {})[d.weekday()] = h

        for wn, wd_hours in by_week.items():
            week_data[wn].append({
                "wbs": p.get("wbs", ""),
                "description": p.get("description", p_name),
                "aa_type": p.get("aa_type", ""),
                "memo": p.get("memo", ""),
                "hours": wd_hours,  # {0: h, 1: h, ...}  0=Mon
            })

    return week_data


def write_catw_excel(excel_path: str, sheet_name: str, week_data: dict, year: int, month: int) -> None:
    """week_data を Excel の CATW シートに書き込む。"""
    wb = load_workbook(excel_path, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        print(f"[エラー] シート '{sheet_name}' が見つかりません（存在するシート: {wb.sheetnames}）")
        sys.exit(1)

    ws = wb[sheet_name]

    # 年月を書き込む
    ws.cell(row=4, column=3, value=year)
    ws.cell(row=5, column=3, value=month)

    write_count = 0
    for wn, entries in week_data.items():
        header_row = WEEK_START_ROWS.get(wn)
        if header_row is None:
            continue
        data_start = header_row + 1

        # 7行クリア（G列の数式は触らない）
        for i in range(DATA_ROW_COUNT):
            r = data_start + i
            for col in [COL_WBS, COL_DESCRIPTION, COL_AA_TYPE, COL_MEMO] + list(WEEKDAY_TO_COL.values()):
                ws.cell(row=r, column=col, value=None)

        # データ書き込み
        for i, entry in enumerate(entries):
            if i >= DATA_ROW_COUNT:
                print(f"[警告] Week {wn} のエントリが {DATA_ROW_COUNT} 行を超えています（切り捨て）")
                break
            r = data_start + i
            ws.cell(row=r, column=COL_WBS, value=entry["wbs"])
            ws.cell(row=r, column=COL_DESCRIPTION, value=entry["description"])
            ws.cell(row=r, column=COL_AA_TYPE, value=entry["aa_type"])
            ws.cell(row=r, column=COL_MEMO, value=entry["memo"])
            for weekday, h in entry["hours"].items():
                col = WEEKDAY_TO_COL.get(weekday)
                if col and h > 0:
                    ws.cell(row=r, column=col, value=h)
            write_count += 1

    wb.save(excel_path)
    print(f"✅ CATW Excel に転記しました: {excel_path}（{write_count} 行）")


def main() -> None:
    today = date.today()
    if len(sys.argv) == 3:
        year, month = int(sys.argv[1]), int(sys.argv[2])
    elif len(sys.argv) == 1:
        year, month = today.year, today.month
    else:
        print("使い方: python scripts/transfer_catw.py [年] [月]")
        sys.exit(1)

    config = load_config()
    projects = config.get("projects", [])
    if not projects:
        print("[エラー] config.yaml に projects が定義されていません")
        sys.exit(1)

    md_path = Path("勤怠") / str(year) / f"{month}月.md"
    if not md_path.exists():
        print(f"[エラー] MDファイルが見つかりません: {md_path}")
        sys.exit(1)

    md_text = md_path.read_text(encoding="utf-8")
    project_hours = parse_catw_table(md_text, year, month, projects)

    total = sum(len(v) for v in project_hours.values())
    if total == 0:
        print("[情報] CATW テーブルに工数データがありません。転記をスキップします。")
        sys.exit(0)

    week_data = build_week_data(project_hours, projects)

    catw_cfg = config.get("catw", {}).get("excel", {})
    excel_path = catw_cfg.get("path", "")
    sheet_name = catw_cfg.get("sheet_name", "CATW")

    if not excel_path or "/path/to/" in excel_path:
        print("[エラー] config.yaml の catw.excel.path を実際のExcelファイルパスに設定してください")
        sys.exit(1)

    if not Path(excel_path).exists():
        print(f"[エラー] Excelファイルが見つかりません: {excel_path}")
        sys.exit(1)

    print(f"\n転記対象: {year}年{month}月 → {excel_path}（{sheet_name} シート）")
    for p_name, dh in project_hours.items():
        if dh:
            total_h = sum(dh.values())
            print(f"  {p_name}: {len(dh)}日 / {total_h}h")

    ans = input("\n転記を実行しますか？ [y/N]: ").strip().lower()
    if ans != "y":
        print("キャンセルしました。")
        sys.exit(2)  # 2 = ユーザーキャンセル（0=成功、1=エラーと区別）

    write_catw_excel(excel_path, sheet_name, week_data, year, month)


if __name__ == "__main__":
    main()
