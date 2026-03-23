# -*- coding: utf-8 -*-
"""
CATW勤怠システム Playwright自動化スクリプト
WebDriver不要版 - ブラウザ更新時のメンテナンス不要
"""

import sys
import os
import json
import time
import traceback
from pathlib import Path
from datetime import datetime, timedelta
import re
import logging

# ログ設定（スクリプトと同じフォルダにログファイルを出力）
_log_dir = os.path.dirname(os.path.abspath(__file__))
_log_file = os.path.join(_log_dir, "catw_debug.log")
logging.basicConfig(
    filename=_log_file,
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    encoding="utf-8",
)
_logger = logging.getLogger("catw")

# printをログにも出力するラッパー
_original_print = print
def print(*args, **kwargs):
    _original_print(*args, **kwargs)
    try:
        msg = " ".join(str(a) for a in args)
        _logger.info(msg)
    except:
        pass

def _log_environment_info():
    """起動時に環境情報をログに記録（トラブルシュート用）"""
    import importlib.metadata
    info_lines = [
        "=" * 60,
        "[環境情報]",
        f"  Python: {sys.version}",
        f"  実行ファイル: {sys.executable}",
        f"  スクリプト: {os.path.abspath(__file__)}",
        f"  作業ディレクトリ: {os.getcwd()}",
        f"  OS: {os.name} / {sys.platform}",
        f"  引数: {sys.argv}",
    ]
    # ライブラリバージョン
    for pkg in ["playwright", "openpyxl", "pywin32"]:
        try:
            ver = importlib.metadata.version(pkg)
            info_lines.append(f"  {pkg}: {ver}")
        except importlib.metadata.PackageNotFoundError:
            info_lines.append(f"  {pkg}: 未インストール")
    # Edge検出
    edge_candidates = [
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    ]
    local_app = os.environ.get("LOCALAPPDATA", "")
    if local_app:
        edge_candidates.append(os.path.join(local_app, r"Microsoft\Edge\Application\msedge.exe"))
    edge_found = [p for p in edge_candidates if os.path.isfile(p)]
    info_lines.append(f"  Edge: {edge_found[0] if edge_found else '検出されず'}")
    info_lines.append(f"  LOCALAPPDATA: {local_app}")
    info_lines.append("=" * 60)
    for line in info_lines:
        _logger.info(line)

_log_environment_info()

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
except ImportError:
    print("="*60)
    print("エラー: Playwrightがインストールされていません")
    print("="*60)
    print("【原因】pip install playwright が実行されていません。")
    print("【対処法】コマンドプロンプトで以下を実行してください:")
    print("  pip install playwright")
    print("")
    print(f"  現在のPython: {sys.executable}")
    print(f"  バージョン: {sys.version}")
    _logger.error(f"playwright未インストール: {traceback.format_exc()}")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: openpyxlがインストールされていません。Excel連携機能は使用できません。")
    print("【対処法】コマンドプロンプトで以下を実行: pip install openpyxl")
    _logger.warning(f"openpyxl未インストール: {traceback.format_exc()}")

# win32com（開いているExcelに書き込むため）
try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False
    print("警告: win32comがインストールされていません。開いているExcelへの書き込みができません。")
    print("【対処法】コマンドプロンプトで以下を実行: pip install pywin32")
    _logger.warning(f"win32com未インストール: {traceback.format_exc()}")

# デフォルトのExcelファイルパス
# DEFAULT_EXCEL_PATH is not used when called from Excel VBA (path is passed as argument)
DEFAULT_EXCEL_PATH = ""

# デバッグポート（既存ブラウザに接続するため）
DEBUG_PORT = 9222

# Excelファイルの構造定義
EXCEL_SHEET_NAME = "CATW"
WEEK_START_ROWS = {
    1: 7, 2: 19, 3: 31, 4: 43, 5: 55, 6: 67,
}
DATA_ROW_COUNT = 7

# 列定義
COL_WBS = 3          # C列: WBS Element
COL_DESCRIPTION = 4  # D列: Description
COL_AA_TYPE = 5      # E列: Attend/Absence Type
COL_MON = 8          # H列: 月曜
COL_TUE = 9          # I列: 火曜
COL_WED = 10         # J列: 水曜
COL_THU = 11         # K列: 木曜
COL_FRI = 12         # L列: 金曜
COL_SAT = 13         # M列: 土曜
COL_SUN = 14         # N列: 日曜

def extract_month_from_filename(filename):
    """ファイル名から月を抽出（後方互換用）"""
    if not filename:
        return None
    match = re.search(r'FY\d{2}\s*(\d{1,2})月', filename)
    if match:
        return int(match.group(1))
    return None


def get_year_month_from_excel():
    """ExcelのC4(年)/C5(月)セルから年月を取得。
    
    Returns:
        (year, month) または (None, None)
    """
    if not WIN32COM_AVAILABLE:
        return None, None
    try:
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception as e:
            _logger.debug(f"Excelインスタンス取得失敗（Excelが開いていない可能性）: {e}")
            return None, None
        
        target_wb = None
        for i in range(1, excel.Workbooks.Count + 1):
            wb = excel.Workbooks.Item(i)
            if "CATW" in wb.Name:
                target_wb = wb
                break
        
        if not target_wb:
            return None, None
        
        ws = target_wb.Sheets(EXCEL_SHEET_NAME)
        year_val = ws.Cells(4, 3).Value   # C4 = Year
        month_val = ws.Cells(5, 3).Value  # C5 = Month
        
        year = int(year_val) if year_val else None
        month = int(month_val) if month_val else None
        
        if year and month:
            print(f"  Excelセルから年月を取得: {year}年{month}月")
            return year, month
        return None, None
    except Exception as e:
        _logger.error(f"年月取得エラー: {e}\n{traceback.format_exc()}")
        print(f"  年月取得エラー: {e}")
        print("  【対処法】CATWマクロExcelが開いているか確認してください")
        return None, None


def get_target_month_days(week_num, year=None, month=None):
    """指定された週で対象月に属する曜日のセットを返す。
    
    例: 2月 Week1 が 1/26(月)~2/1(日) の場合 → {"sun"} のみ2月
    境界でない週は全曜日が返る。
    """
    if month is None:
        month = datetime.now().month
    if year is None:
        year = datetime.now().year
    
    first_day = datetime(year, month, 1)
    first_weekday = first_day.weekday()
    
    if first_weekday == 0:
        week1_monday = first_day
    else:
        week1_monday = first_day - timedelta(days=first_weekday)
    
    target_monday = week1_monday + timedelta(weeks=week_num - 1)
    
    day_keys = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
    result = set()
    for i, key in enumerate(day_keys):
        day_date = target_monday + timedelta(days=i)
        if day_date.month == month:
            result.add(key)
    
    return result


def get_week_date(week_num=None, month=None, year=None):
    """週番号から対象週の日付を取得
    
    year/month が指定されればそれを使い、datetime.now() に依存しない。
    """
    if month is None:
        today = datetime.now()
    else:
        if year is None:
            year = datetime.now().year
        today = datetime(year, month, 1)
    
    if week_num is None:
        return today
    
    first_day = today.replace(day=1)
    first_weekday = first_day.weekday()
    
    if first_weekday == 0:
        week1_monday = first_day
    else:
        week1_monday = first_day - timedelta(days=first_weekday)
    
    target_monday = week1_monday + timedelta(weeks=week_num - 1)
    target_date = target_monday + timedelta(days=2)  # 水曜日
    
    return target_date


def get_current_week_num():
    """今日の日付から現在の週番号を取得"""
    today = datetime.now()
    first_day = today.replace(day=1)
    first_weekday = first_day.weekday()
    
    if first_weekday == 0:
        week1_monday = first_day
    else:
        week1_monday = first_day - timedelta(days=first_weekday)
    
    days_from_week1 = (today - week1_monday).days
    week_num = (days_from_week1 // 7) + 1
    
    return max(1, min(6, week_num))


class CATWAutomation:
    """CATW勤怠システム自動化クラス（Playwright版）"""
    
    CATW_URL = "https://pj1-ent.sapnet.hpecorp.net/hps-ic"
    
    def __init__(self):
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
    
    def connect(self, reuse_browser=False):
        """Edgeブラウザに接続"""
        try:
            self.playwright = sync_playwright().start()
            
            if reuse_browser:
                # 既存のブラウザに接続
                print("既存のブラウザに接続中...")
                try:
                    self.browser = self.playwright.chromium.connect_over_cdp(
                        f"http://127.0.0.1:{DEBUG_PORT}"
                    )
                    contexts = self.browser.contexts
                    if contexts:
                        self.context = contexts[0]
                        pages = self.context.pages
                        
                        # CATWタブを探す
                        for p in pages:
                            if 'HPE-IC' in p.title() or 'sapnet' in p.url:
                                self.page = p
                                print(f"CATWタブに接続: {p.title()}")
                                return True
                        
                        # CATWタブがなければ最初のページを使用
                        if pages:
                            self.page = pages[0]
                            print(f"現在のページ: {self.page.title()}")
                            return True
                    
                    print("ブラウザにページがありません")
                    return False
                    
                except Exception as e:
                    _logger.error(f"既存ブラウザ接続失敗: {e}\n{traceback.format_exc()}")
                    print(f"既存ブラウザへの接続に失敗: {e}")
                    print("\n【対処法】")
                    print("  1. 「Open CATW」ボタンでブラウザを起動してください")
                    print("  2. 既にEdgeが開いている場合、一度すべてのEdgeを閉じてから再実行")
                    print("  3. タスクマネージャーで msedge.exe が残っていないか確認")
                    return False
            else:
                # 新規ブラウザを起動
                return self.start_debug_browser()
                
        except Exception as e:
            _logger.error(f"Playwright初期化エラー: {e}\n{traceback.format_exc()}")
            print(f"Playwright初期化エラー: {e}")
            print("\n【対処法】")
            print("  1. pip install playwright を実行してください")
            print("  2. Pythonのバージョンが3.9以上か確認: python --version")
            print(f"  現在: {sys.executable} ({sys.version})")
            return False
    
    @staticmethod
    def _find_edge_path():
        """複数の方法でMicrosoft Edgeの実行パスを検出する"""
        import shutil

        # 1. よくあるインストールパスを確認
        candidate_paths = [
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        ]
        # ユーザー別インストール（AppData）
        local_app = os.environ.get("LOCALAPPDATA", "")
        if local_app:
            candidate_paths.append(os.path.join(local_app, r"Microsoft\Edge\Application\msedge.exe"))

        for p in candidate_paths:
            if os.path.isfile(p):
                return p

        # 2. PATH上の msedge / msedge.exe を探す
        found = shutil.which("msedge")
        if found:
            return found

        # 3. Windowsレジストリから取得
        try:
            import winreg
            reg_paths = [
                (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\msedge.exe"),
                (winreg.HKEY_CURRENT_USER,  r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\msedge.exe"),
            ]
            for hive, sub_key in reg_paths:
                try:
                    with winreg.OpenKey(hive, sub_key) as key:
                        val, _ = winreg.QueryValueEx(key, "")
                        if val and os.path.isfile(val):
                            return val
                except FileNotFoundError:
                    continue
        except ImportError:
            pass

        return None

    def start_debug_browser(self):
        """デバッグポート付きでEdgeを新しいウィンドウで起動"""
        import subprocess
        
        # Playwrightが初期化されていなければ初期化
        if not self.playwright:
            self.playwright = sync_playwright().start()
        
        edge_path = self._find_edge_path()
        if not edge_path:
            print("エラー: Microsoft Edgeが見つかりません。")
            print("以下を確認してください:")
            print("  1. Microsoft Edgeがインストールされているか")
            print("  2. 標準の場所にインストールされているか")
            print("     (C:\\Program Files\\Microsoft\\Edge\\Application\\msedge.exe)")
            return False
        
        print(f"Edge検出パス: {edge_path}")
        
        # 独立したプロファイルで起動（既存Edgeが開いていてもデバッグポートが有効になる）
        user_data_dir = os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~')), 'CATW_Edge_Profile')
        cmd = [edge_path, f"--remote-debugging-port={DEBUG_PORT}", f"--user-data-dir={user_data_dir}", self.CATW_URL]
        
        # プロファイルが存在するか確認
        if os.path.exists(user_data_dir):
            print("Edgeを起動中（CATW専用プロファイル - ログイン情報は保存済み）...")
        else:
            print("Edgeを起動中（CATW専用プロファイル - 初回のみログインが必要です）...")
        
        subprocess.Popen(cmd)
        
        # 接続リトライ（最大3回、各3秒待ち）
        for attempt in range(1, 4):
            time.sleep(3)
            try:
                self.browser = self.playwright.chromium.connect_over_cdp(
                    f"http://127.0.0.1:{DEBUG_PORT}"
                )
                contexts = self.browser.contexts
                if contexts:
                    self.context = contexts[0]
                    if self.context.pages:
                        self.page = self.context.pages[0]
                        print("ブラウザに接続しました")
                        return True
            except Exception as e:
                if attempt < 3:
                    print(f"ブラウザ接続を再試行中... ({attempt}/3) - {e}")
                else:
                    _logger.error(f"ブラウザ接続エラー(3回失敗): {e}\n{traceback.format_exc()}")
                    print(f"ブラウザ接続エラー: {e}")
                    print("\n【対処法】")
                    print("  1. すべてのEdgeを閉じてから再実行してください")
                    print("  2. タスクマネージャーで msedge.exe をすべて終了")
                    print(f"  3. ポート {DEBUG_PORT} が他プログラムに使われていないか確認")
                    print(f"  4. ファイアウォールがポート {DEBUG_PORT} をブロックしていないか確認")
        
        return False
    
    def _get_right_frame(self):
        """rightフレームを取得"""
        try:
            frames = self.page.frames
            for frame in frames:
                if frame.name == "right":
                    return frame
                # フレーム名がなくてもURLで判定
                if "right" in frame.url.lower() or "timeentry" in frame.url.lower():
                    return frame
            # フレーム名一覧を表示（デバッグ用）
            print(f"  利用可能なフレーム: {[f.name or f.url[:50] for f in frames]}")
        except Exception as e:
            _logger.error(f"フレーム取得エラー: {e}\n{traceback.format_exc()}")
            print(f"  フレーム取得エラー: {e}")
            print("  【対処法】CATWにTime Entry画面が表示されているか確認してください")
        return None
    
    def navigate_to_week(self, week_num=None, month=None, year=None):
        """指定された週に移動"""
        target_date = get_week_date(week_num, month, year)
        date_str = target_date.strftime("%m/%d/%Y")
        
        print(f"対象週: Week {week_num if week_num else '(自動)'} → 日付: {date_str}")
        
        try:
            right_frame = self._get_right_frame()
            
            if not right_frame:
                print("rightフレームが見つかりません")
                return False
            
            # 日付入力フィールドを待機
            date_input = right_frame.locator("#lo_sel_date")
            date_input.wait_for(timeout=10000)
            date_input.fill(date_str)
            
            # Dateボタンをクリック
            date_button = right_frame.locator("#DateButton")
            date_button.click()
            
            print(f"  週を切り替え中... (日付: {date_str})")
            
            # テーブル読み込み完了を待機
            time.sleep(3)
            try:
                # input要素が表示されるまで待機（最大10秒）
                right_frame.locator("input[id^='data_tab_']").first.wait_for(state="visible", timeout=10000)
                print("  テーブル読み込み完了")
            except Exception as e:
                _logger.debug(f"テーブル読み込み待機タイムアウト: {e}")
                print("  警告: テーブルの読み込みに時間がかかっています")
                time.sleep(3)  # 追加待機
            
            return True
            
        except Exception as e:
            _logger.error(f"週切替失敗: {e}\n{traceback.format_exc()}")
            print(f"週切り替えに失敗: {e}")
            print("\n【対処法】")
            print("  1. CATWにTime Entry画面が表示されているか確認")
            print("  2. ブラウザがフリーズしている場合、再起動してください")
            return False
    
    def read_week_data(self):
        """現在表示されている週のデータを読み取る"""
        data = []
        
        try:
            right_frame = self._get_right_frame()
            
            if not right_frame:
                print("  rightフレームが見つかりません")
                return data
            
            # テーブル読み込み待機
            print("  テーブル検索中...")
            time.sleep(2)
            
            # テーブル番号を検出（リトライ付き）
            table_num = None
            for retry in range(5):
                try:
                    # input要素が表示されるまで待機
                    right_frame.locator("input[id^='data_tab_']").first.wait_for(state="visible", timeout=3000)
                    inputs = right_frame.locator("input[id^='data_tab_']").all()
                    print(f"  入力要素数: {len(inputs)}")
                    table_nums = set()
                    
                    for inp in inputs:
                        try:
                            inp_id = inp.get_attribute("id", timeout=1000)
                            if inp_id:
                                parts = inp_id.split("_")
                                if len(parts) >= 4 and parts[2].isdigit():
                                    table_nums.add(parts[2])
                        except Exception:
                            pass
                    
                    if table_nums:
                        table_num = max(table_nums)  # 最新のテーブル番号を使用
                        break
                except Exception as e:
                    _logger.debug(f"テーブル検出リトライ({retry+1}/5): {e}")
                    pass
                time.sleep(1)
            
            if not table_num:
                print("  警告: data_tab要素が見つかりません")
                return data
            
            print(f"  テーブル番号: {table_num}")
            
            # 行番号を収集
            inputs = right_frame.locator("input[id^='data_tab_']").all()
            row_numbers = set()
            for inp in inputs:
                inp_id = inp.get_attribute("id")
                prefix = f"data_tab_{table_num}_"
                if inp_id and inp_id.startswith(prefix) and inp_id.endswith("_1"):
                    parts = inp_id.split("_")
                    if len(parts) >= 5:
                        try:
                            row_num = int(parts[3])
                            row_numbers.add(row_num)
                        except ValueError:
                            pass
            
            print(f"  検出された行: {sorted(row_numbers)}")
            
            if not row_numbers:
                print("  データ行がありません")
                return data
            
            # 各行のデータを読み取る
            for row_num in sorted(row_numbers):
                print(f"  行 {row_num} を読み取り中...", end="", flush=True)
                row_data = {
                    "row": row_num,
                    "wbs_element": "",
                    "description": "",
                    "aa_type": "",
                    "aa_type_code": "",
                    "hours": {
                        "mon": 0.0, "tue": 0.0, "wed": 0.0,
                        "thu": 0.0, "fri": 0.0, "sat": 0.0, "sun": 0.0
                    }
                }
                
                base_id = f"data_tab_{table_num}_{row_num}"
                
                # WBS Element (_1)
                try:
                    wbs_elem = right_frame.locator(f"#{base_id}_1")
                    row_data["wbs_element"] = wbs_elem.input_value(timeout=3000) or ""
                except Exception as e:
                    pass

                # Description - テーブル構造のTD[2]から取得（_2は存在しない）
                try:
                    wbs_elem = right_frame.locator(f"#{base_id}_1")
                    desc_text = wbs_elem.evaluate("""el => {
                        // WBS入力から親テーブルを2レベル上に辿る
                        let table = el.closest('table');
                        if (table) table = table.parentElement.closest('table');
                        if (!table) return '';

                        // 同じ行を探す
                        let rows = table.querySelectorAll(':scope > tbody > tr');
                        for (let row of rows) {
                            if (row.innerHTML.includes(el.id)) {
                                let tds = row.querySelectorAll(':scope > td');
                                // TD[2]がDescription列
                                if (tds.length >= 3) {
                                    return tds[2].innerText.trim();
                                }
                            }
                        }
                        return '';
                    }""")
                    row_data["description"] = desc_text or ""
                except Exception as e:
                    pass

                # Attend/Absence Type (_3) - Select要素
                try:
                    aa_select = right_frame.locator(f"#{base_id}_3")
                    row_data["aa_type"] = aa_select.evaluate("el => el.options[el.selectedIndex]?.text || ''")
                    row_data["aa_type_code"] = aa_select.input_value(timeout=3000) or ""
                except Exception as e:
                    _logger.debug(f"行{row_num} AA Type取得スキップ: {e}")
                
                # 時間データ - 正しい列マッピング
                # _6=月, _7=火, _8=水, _9=木, _10=金, _11=土, _12=日
                day_map = {6: "mon", 7: "tue", 8: "wed", 9: "thu", 10: "fri", 11: "sat", 12: "sun"}
                for col, day in day_map.items():
                    try:
                        hour_input = right_frame.locator(f"#{base_id}_{col}")
                        value = hour_input.input_value(timeout=2000) or "0"
                        row_data["hours"][day] = float(value) if value else 0.0
                    except Exception as e:
                        _logger.debug(f"行{row_num} {day}時間取得スキップ: {e}")
                
                # データがある行のみ追加
                if row_data["wbs_element"] or any(v > 0 for v in row_data["hours"].values()):
                    data.append(row_data)
                    print(f" WBS: {row_data['wbs_element'][:15]}...")
                else:
                    print(" (空)")
            
            return data
            
        except Exception as e:
            _logger.error(f"データ読み取りエラー: {e}\n{traceback.format_exc()}")
            print(f"データ読み取りエラー: {e}")
            print("\n【対処法】")
            print("  1. CATWのTime Entryテーブルが表示されているか確認")
            print("  2. ページが完全に読み込まれてから再実行")
            return data
    
    def input_week_data(self, data, week_num=None, month=None, year=None):
        """
        CATWにデータを入力
        
        ■ 基本方針: 前月データは絶対に消さない・触らない
        
        【非境界週】全曜日が当月 → 既存行を全削除して Excel データを入力
        【境界週】  前月の曜日がある → 既存行は一切削除しない
          - 既存行のWBSとExcelのWBSが一致 → 当月曜日の列だけ上書き
          - Excelにあるが既存行にない → 新規行を追加し当月曜日だけ入力
          - 既存行にあるがExcelにない → 完全ノータッチ（前月データ保護）
        
        列: _1=WBS, _3=AA Type, _6=月, _7=火, _8=水, _9=木, _10=金, _11=土, _12=日
        """
        try:
            right_frame = self._get_right_frame()
            
            if not right_frame:
                print("rightフレームが見つかりません")
                return False
            
            # 対象月の曜日を算出
            if month is None:
                month = datetime.now().month
            if year is None:
                year = datetime.now().year
            if week_num:
                target_days = get_target_month_days(week_num, year=year, month=month)
            else:
                target_days = {"mon", "tue", "wed", "thu", "fri", "sat", "sun"}
            
            all_days = {"mon", "tue", "wed", "thu", "fri", "sat", "sun"}
            non_target_days = all_days - target_days
            is_boundary_week = len(non_target_days) > 0
            
            if is_boundary_week:
                print(f"  ※ 月境界週: {', '.join(sorted(non_target_days))} は前月 → ノータッチ")
            print(f"  対象曜日（当月）: {', '.join(sorted(target_days))}")
            
            print("  テーブル読み込み待機中...")
            time.sleep(2)
            
            day_map = {"mon": 6, "tue": 7, "wed": 8, "thu": 9, "fri": 10, "sat": 11, "sun": 12}
            
            if is_boundary_week:
                # ============================================================
                # 【境界週】既存行を削除しない。当月曜日だけ上書きする。
                # ============================================================
                return self._input_boundary_week(
                    right_frame, data, target_days, day_map)
            else:
                # ============================================================
                # 【非境界週】全曜日が当月。既存行を全削除→再入力。
                # ============================================================
                return self._input_full_week(
                    right_frame, data, target_days)
            
        except Exception as e:
            _logger.error(f"データ入力エラー: {e}\n{traceback.format_exc()}")
            print(f"データ入力エラー: {e}")
            print("\n【対処法】")
            print("  1. CATWのTime Entryテーブルが表示されているか確認")
            print("  2. ページが完全に読み込まれてから再実行")
            return False
    
    def _input_boundary_week(self, right_frame, data, target_days, day_map):
        """境界週の入力: 既存行を削除せず、当月曜日の列だけ上書き"""
        
        # --- Step 1: 既存行の WBS マップを作成 & 空行を把握 ---
        existing_rows = self._find_all_input_rows(right_frame)
        wbs_to_pos = {}  # WBS文字列 → 位置インデックス
        empty_positions = []  # WBSが空の行（使える空行）
        
        print(f"  既存行: {len(existing_rows)}行（削除しません）")
        for pos_idx, (table_num, row_num) in enumerate(existing_rows):
            base_id = f"data_tab_{table_num}_{row_num}"
            try:
                wbs_input = right_frame.locator(f"#{base_id}_1")
                wbs_val = (wbs_input.input_value(timeout=2000) or "").strip()
            except Exception as e:
                _logger.debug(f"境界週 WBS取得失敗(行{row_num}): {e}")
                wbs_val = ""
            if wbs_val:
                wbs_to_pos[wbs_val] = pos_idx
                print(f"    既存行[{pos_idx}] WBS={wbs_val[:30]}")
            else:
                empty_positions.append(pos_idx)
                print(f"    既存行[{pos_idx}] (空行 - 新規入力に使用可能)")
        
        # --- Step 2: Excelデータを処理 ---
        matched_wbs = set()
        new_rows_data = []  # 既存行にないExcelデータ
        
        for row_data in data:
            wbs = row_data.get("wbs_element", "").strip()
            if not wbs:
                continue
            
            if wbs in wbs_to_pos:
                # 既存行にWBSが一致 → 当月曜日の列だけ上書き
                pos_idx = wbs_to_pos[wbs]
                matched_wbs.add(wbs)
                
                # 最新の行情報を再取得
                current_rows = self._find_all_input_rows(right_frame)
                if pos_idx >= len(current_rows):
                    print(f"  [スキップ] WBS={wbs[:25]} 行が見つかりません")
                    continue
                
                table_num, row_num = current_rows[pos_idx]
                base_id = f"data_tab_{table_num}_{row_num}"
                print(f"\n  [上書き] {base_id} WBS={wbs[:25]} （当月曜日のみ）")
                
                self._fill_row_hours(right_frame, base_id, row_data, target_days=target_days)
                time.sleep(0.3)
            else:
                # 既存行にない → 新規行として追加
                new_rows_data.append(row_data)
        
        # 既存行でExcelにないもの → 前月曜日はノータッチ、当月曜日はクリア
        untouched = set(wbs_to_pos.keys()) - matched_wbs
        if untouched:
            print(f"\n  [当月クリア] Excelにない既存行の当月曜日を0にします:")
            for wbs in untouched:
                pos_idx = wbs_to_pos[wbs]
                current_rows = self._find_all_input_rows(right_frame)
                if pos_idx >= len(current_rows):
                    continue
                table_num, row_num = current_rows[pos_idx]
                base_id = f"data_tab_{table_num}_{row_num}"
                print(f"    WBS={wbs[:30]} → 当月曜日をクリア")
                # 空のhoursデータで当月曜日をクリア
                empty_row = {"hours": {}}
                self._fill_row_hours(right_frame, base_id, empty_row, target_days=target_days)
                time.sleep(0.3)
        
        # --- Step 3: 新規行を入力（空行を先に使い、足りなければAddLine） ---
        if new_rows_data:
            print(f"\n  新規行入力: {len(new_rows_data)}行 (空行: {len(empty_positions)}個)")
            for row_data in new_rows_data:
                if empty_positions:
                    # 既存の空行を使う
                    pos_idx = empty_positions.pop(0)
                    right_frame = self._fill_existing_row(
                        right_frame, pos_idx, row_data, target_days)
                    if not right_frame:
                        print("  フレーム再取得失敗、入力中断")
                        return False
                else:
                    # 空行がない → AddLineで追加
                    right_frame = self._add_and_fill_row(
                        right_frame, row_data, target_days)
                    if not right_frame:
                        print("  フレーム再取得失敗、追加中断")
                        return False
        
        print("\n  境界週入力完了")
        return True
    
    def _fill_existing_row(self, right_frame, pos_idx, row_data, target_days):
        """既存の空行に WBS + AA Type + 当月曜日を入力。
        戻り値: 更新後の right_frame (失敗時 None)
        """
        wbs = row_data.get("wbs_element", "")[:25]
        
        current_rows = self._find_all_input_rows(right_frame)
        if pos_idx >= len(current_rows):
            print(f"  [スキップ] 空行[{pos_idx}]が見つかりません")
            return right_frame
        
        table_num, row_num = current_rows[pos_idx]
        base_id = f"data_tab_{table_num}_{row_num}"
        print(f"\n  [空行使用] {base_id} WBS={wbs}")
        
        # WBS入力
        if row_data.get("wbs_element"):
            try:
                wbs_input = right_frame.locator(f"#{base_id}_1")
                wbs_input.click()
                time.sleep(0.1)
                wbs_input.fill(row_data["wbs_element"])
                wbs_input.press("Tab")
                time.sleep(0.5)
                right_frame = self._get_right_frame()
                if not right_frame:
                    return None
                print(f"        WBS入力OK")
            except Exception as e:
                print(f"        WBS入力エラー: {e}")
        
        # テーブル番号再取得
        current_rows = self._find_all_input_rows(right_frame)
        if pos_idx < len(current_rows):
            table_num, row_num = current_rows[pos_idx]
            base_id = f"data_tab_{table_num}_{row_num}"
        
        # AA Type入力
        if row_data.get("aa_type_code"):
            try:
                aa_select = right_frame.locator(f"#{base_id}_3")
                aa_select.click()
                time.sleep(0.1)
                aa_select.select_option(value=row_data["aa_type_code"])
                time.sleep(0.2)
                print(f"        AA Type入力OK: {row_data['aa_type_code']}")
            except Exception as e:
                print(f"        AA Type選択エラー: {e}")
        
        # 当月曜日だけ入力
        self._fill_row_hours(right_frame, base_id, row_data, target_days=target_days)
        time.sleep(0.3)
        
        return right_frame
    
    def _input_full_week(self, right_frame, data, target_days):
        """非境界週の入力: 全削除→再入力"""
        
        existing_rows = self._find_all_input_rows(right_frame)
        
        # --- 既存行を全削除 ---
        if existing_rows:
            print(f"  既存{len(existing_rows)}行を削除中...")
            self._delete_all_input_rows(right_frame)
            time.sleep(1)
            right_frame = self._get_right_frame()
            if not right_frame:
                print("  フレーム再取得失敗")
                return False
        
        # --- Excelデータを全行入力 ---
        print(f"\n  入力開始: {len(data)}行")
        
        for idx, row_data in enumerate(data):
            # 毎回行一覧を再取得（テーブル番号変更に対応）
            current_rows = self._find_all_input_rows(right_frame)
            
            # 必要に応じてAddLineで行を追加
            add_retries = 0
            while idx >= len(current_rows) and add_retries < 5:
                try:
                    add_btn = right_frame.locator("#AddLine")
                    add_btn.click()
                    time.sleep(1)
                    right_frame = self._get_right_frame()
                    if not right_frame:
                        print("  フレーム再取得失敗")
                        break
                    current_rows = self._find_all_input_rows(right_frame)
                    add_retries += 1
                except Exception as e:
                    print(f"  Add Lineエラー: {e}")
                    break
            
            if not right_frame or idx >= len(current_rows):
                print(f"  行追加に失敗 (必要: {idx+1}, 利用可能: {len(current_rows)})")
                break
            
            table_num, row_num = current_rows[idx]
            base_id = f"data_tab_{table_num}_{row_num}"
            wbs = row_data.get("wbs_element", "")[:25]
            print(f"\n  [入力 {idx+1}/{len(data)}] {base_id} WBS={wbs}")
            
            # WBS入力
            if row_data.get("wbs_element"):
                try:
                    wbs_input = right_frame.locator(f"#{base_id}_1")
                    wbs_input.click()
                    time.sleep(0.1)
                    wbs_input.fill(row_data["wbs_element"])
                    wbs_input.press("Tab")
                    time.sleep(0.5)
                    right_frame = self._get_right_frame()
                    if not right_frame:
                        print("        フレーム再取得失敗")
                        break
                    print(f"        WBS入力OK")
                except Exception as e:
                    print(f"        WBS入力エラー: {e}")
            
            # テーブル番号が変わっている可能性があるため再検索
            current_rows = self._find_all_input_rows(right_frame)
            if idx < len(current_rows):
                table_num, row_num = current_rows[idx]
                base_id = f"data_tab_{table_num}_{row_num}"
            
            # AA Type入力
            if row_data.get("aa_type_code"):
                try:
                    aa_select = right_frame.locator(f"#{base_id}_3")
                    aa_select.click()
                    time.sleep(0.1)
                    aa_select.select_option(value=row_data["aa_type_code"])
                    time.sleep(0.2)
                    print(f"        AA Type入力OK: {row_data['aa_type_code']}")
                except Exception as e:
                    print(f"        AA Type選択エラー: {e}")
            
            # 時間データ（全曜日）
            self._fill_row_hours(right_frame, base_id, row_data, target_days=target_days)
            time.sleep(0.3)
        
        return True
    
    def _add_and_fill_row(self, right_frame, row_data, target_days):
        """新規行を AddLine で追加し、WBS + AA Type + 当月曜日のみ入力。
        戻り値: 更新後の right_frame (失敗時 None)
        """
        wbs = row_data.get("wbs_element", "")[:25]
        
        # AddLine前の行数を記録
        before_rows = self._find_all_input_rows(right_frame)
        before_count = len(before_rows)
        
        # AddLine
        try:
            add_btn = right_frame.locator("#AddLine")
            add_btn.click()
            time.sleep(1)
            right_frame = self._get_right_frame()
            if not right_frame:
                return None
        except Exception as e:
            print(f"  Add Lineエラー: {e}")
            return right_frame
        
        # 新しい行を取得（末尾の行）
        current_rows = self._find_all_input_rows(right_frame)
        if len(current_rows) <= before_count:
            print(f"  [追加失敗] WBS={wbs} 行が増えませんでした")
            return right_frame
        
        new_idx = len(current_rows) - 1
        table_num, row_num = current_rows[new_idx]
        base_id = f"data_tab_{table_num}_{row_num}"
        print(f"\n  [新規追加] {base_id} WBS={wbs}")
        
        # WBS入力
        if row_data.get("wbs_element"):
            try:
                wbs_input = right_frame.locator(f"#{base_id}_1")
                wbs_input.click()
                time.sleep(0.1)
                wbs_input.fill(row_data["wbs_element"])
                wbs_input.press("Tab")
                time.sleep(0.5)
                right_frame = self._get_right_frame()
                if not right_frame:
                    return None
                print(f"        WBS入力OK")
            except Exception as e:
                print(f"        WBS入力エラー: {e}")
        
        # テーブル番号再取得
        current_rows = self._find_all_input_rows(right_frame)
        if new_idx < len(current_rows):
            table_num, row_num = current_rows[new_idx]
            base_id = f"data_tab_{table_num}_{row_num}"
        
        # AA Type入力
        if row_data.get("aa_type_code"):
            try:
                aa_select = right_frame.locator(f"#{base_id}_3")
                aa_select.click()
                time.sleep(0.1)
                aa_select.select_option(value=row_data["aa_type_code"])
                time.sleep(0.2)
                print(f"        AA Type入力OK: {row_data['aa_type_code']}")
            except Exception as e:
                print(f"        AA Type選択エラー: {e}")
        
        # 当月曜日だけ入力
        self._fill_row_hours(right_frame, base_id, row_data, target_days=target_days)
        time.sleep(0.3)
        
        return right_frame
    
    def _fill_row_hours(self, frame, base_id, row_data, target_days=None):
        """行の時間データを入力。
        
        target_days が指定された場合:
        - target_days に含まれる曜日はまずクリア（空文字）してから値を入力
        - target_days に含まれない曜日は一切触らない（前月データ保護）
        """
        day_map = {"mon": 6, "tue": 7, "wed": 8, "thu": 9, "fri": 10, "sat": 11, "sun": 12}
        if target_days is None:
            target_days = set(day_map.keys())
        
        hours = row_data.get("hours", {})
        entered_hours = []
        cleared_hours = []
        
        for day, col in day_map.items():
            if day not in target_days:
                continue  # 前月の曜日 → 触らない
            
            value = hours.get(day, 0)
            try:
                hour_input = frame.locator(f"#{base_id}_{col}")
                hour_input.click()
                time.sleep(0.05)
                if value > 0:
                    hour_input.fill(str(value))
                    entered_hours.append(f"{day}={value}")
                else:
                    # 対象月の曜日はクリア（0 or 未入力→空にする）
                    hour_input.fill("")
                    cleared_hours.append(day)
                hour_input.press("Tab")
                time.sleep(0.1)
            except Exception as e:
                print(f"        {day}入力エラー: {e}")
        
        if entered_hours:
            print(f"        時間入力OK: {', '.join(entered_hours)}")
        if cleared_hours:
            print(f"        クリア: {', '.join(cleared_hours)}")
    
    def _delete_all_input_rows(self, frame):
        """全ての入力行をDelete Entryで削除"""
        try:
            # 行選択ボタンを取得
            select_btns = frame.locator('button[onclick*="SapTable_selectRow"]').all()
            
            if not select_btns:
                print("    削除する行がありません")
                return
            
            print(f"    {len(select_btns)}行を削除中...")
            
            # 全ての行を選択
            for btn in select_btns:
                try:
                    btn.click()
                    time.sleep(0.1)
                except Exception:
                    pass
            
            time.sleep(0.3)
            
            # Delete Entryをクリック
            delete_btn = frame.locator("#DeleteButton")
            delete_btn.click()
            time.sleep(1)
            
            print(f"    削除完了")
        except Exception as e:
            _logger.error(f"削除エラー: {e}\n{traceback.format_exc()}")
            print(f"    削除エラー: {e}")
            print("    【対処法】CATWページが読み込み完了してから再実行してください")
    
    def _clear_all_input_rows(self, frame):
        """全ての入力行をクリア（後方互換性用）- Delete Entryを使用"""
        self._delete_all_input_rows(frame)
    
    def _find_all_input_rows(self, frame):
        """入力可能な全行（INPUT要素がある行）を探す"""
        rows = []
        try:
            # WBS列のinput要素を探す（_1で終わる）
            inputs = frame.locator("input[id^='data_tab_'][id$='_1']").all()
            
            for inp in inputs:
                inp_id = inp.get_attribute("id")
                if inp_id:
                    # data_tab_{table}_{row}_1 の形式
                    parts = inp_id.split("_")
                    if len(parts) >= 5:
                        table_num = parts[2]
                        row_num = parts[3]
                        rows.append((table_num, row_num))
            
            return rows
        except Exception as e:
            _logger.error(f"入力行検索エラー: {e}\n{traceback.format_exc()}")
            print(f"    入力行検索エラー: {e}")
            return rows
    
    def _find_input_row(self, frame):
        """入力可能な行（INPUT要素がある行）を探す - 後方互換性用"""
        rows = self._find_all_input_rows(frame)
        return rows[0] if rows else None
    
    def save(self):
        """保存ボタンをクリック"""
        try:
            right_frame = self._get_right_frame()
            
            if not right_frame:
                print("rightフレームが見つかりません")
                return False
            
            # 保存ボタンをクリック（2回）
            save_button = right_frame.locator("#SaveButton")
            save_button.click()
            time.sleep(2)
            
            save_button = right_frame.locator("#SaveButton")
            save_button.click()
            time.sleep(2)
            
            print("  保存完了")
            return True
            
        except Exception as e:
            _logger.error(f"保存エラー: {e}\n{traceback.format_exc()}")
            print(f"保存エラー: {e}")
            print("\n【対処法】CATWのTime Entry画面が表示されているか確認してください")
            return False
    
    def close(self):
        """リソースを解放（ブラウザは閉じない）"""
        if self.playwright:
            self.playwright.stop()


def save_to_catw_excel(data, excel_path=None, week_num=None):
    """CATWマクロExcelにデータを保存（開いているExcelに直接書き込み）"""
    if not week_num:
        week_num = get_current_week_num()
    
    print(f"  Excelに保存中: Week {week_num}")
    
    if not data:
        print("  保存するデータがありません")
        return False
    
    # win32comで開いているExcelに書き込む（推奨）
    if WIN32COM_AVAILABLE:
        try:
            # GetActiveObjectで既存のExcelインスタンスに接続
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
            except Exception as e:
                _logger.debug(f"GetActiveObject失敗(保存): {e}")
                excel = win32com.client.Dispatch("Excel.Application")
            
            # 開いているブックから対象を探す
            target_wb = None
            if excel.Workbooks.Count > 0:
                for i in range(1, excel.Workbooks.Count + 1):
                    wb = excel.Workbooks.Item(i)
                    if "CATW" in wb.Name:
                        target_wb = wb
                        break
            
            if not target_wb:
                print("  警告: CATWブックが開いていません。openpyxlで保存します。")
            else:
                ws = target_wb.Sheets(EXCEL_SHEET_NAME)
                
                header_row = WEEK_START_ROWS.get(week_num, 7)
                data_start_row = header_row + 1
                
                # シート保護を解除
                try:
                    ws.Unprotect()
                except Exception as e:
                    _logger.debug(f"シート保護解除スキップ: {e}")

                # まず7行分クリア（G列=数式は保持）
                for i in range(DATA_ROW_COUNT):
                    excel_row = data_start_row + i
                    # C-F列クリア
                    for col in [COL_WBS, COL_DESCRIPTION, COL_AA_TYPE, 6]:  # 6=Memo(F列)
                        ws.Cells(excel_row, col).Value = ""
                    # H-N列クリア
                    for col in [COL_MON, COL_TUE, COL_WED, COL_THU, COL_FRI, COL_SAT, COL_SUN]:
                        ws.Cells(excel_row, col).Value = ""

                for i, row_data in enumerate(data):
                    if i >= DATA_ROW_COUNT:
                        break
                    
                    excel_row = data_start_row + i
                    
                    ws.Cells(excel_row, COL_WBS).Value = row_data.get("wbs_element", "")
                    ws.Cells(excel_row, COL_DESCRIPTION).Value = row_data.get("description", "")
                    ws.Cells(excel_row, COL_AA_TYPE).Value = row_data.get("aa_type", "")
                    
                    hours = row_data.get("hours", {})
                    ws.Cells(excel_row, COL_MON).Value = hours.get("mon", 0) if hours.get("mon", 0) > 0 else ""
                    ws.Cells(excel_row, COL_TUE).Value = hours.get("tue", 0) if hours.get("tue", 0) > 0 else ""
                    ws.Cells(excel_row, COL_WED).Value = hours.get("wed", 0) if hours.get("wed", 0) > 0 else ""
                    ws.Cells(excel_row, COL_THU).Value = hours.get("thu", 0) if hours.get("thu", 0) > 0 else ""
                    ws.Cells(excel_row, COL_FRI).Value = hours.get("fri", 0) if hours.get("fri", 0) > 0 else ""
                    ws.Cells(excel_row, COL_SAT).Value = hours.get("sat", 0) if hours.get("sat", 0) > 0 else ""
                    ws.Cells(excel_row, COL_SUN).Value = hours.get("sun", 0) if hours.get("sun", 0) > 0 else ""
                
                # シート保護を戻す
                try:
                    ws.Protect(DrawingObjects=True, Contents=True, Scenarios=True)
                except Exception as e:
                    _logger.debug(f"シート保護復元スキップ: {e}")
                
                print(f"  ✅ Excel反映完了: {len(data)}行")
                return True
                
        except Exception as e:
            _logger.error(f"win32com書き込みエラー: {e}\n{traceback.format_exc()}")
            print(f"  win32com書き込みエラー: {e}")
            print("  【対処法】")
            print("    1. CATWマクロExcelが開いているか確認")
            print(f"    2. シート名 '{EXCEL_SHEET_NAME}' が存在するか確認")
            print("  openpyxlで保存を試みます...")
    
    # フォールバック: openpyxlで保存（Excelが閉じている場合）
    if not EXCEL_AVAILABLE:
        print("  Excel保存機能が利用できません（openpyxlが必要）")
        return False
    
    if not excel_path:
        excel_path = DEFAULT_EXCEL_PATH
    
    if not os.path.exists(excel_path):
        print(f"  Excelファイルが見つかりません: {excel_path}")
        return False
    
    try:
        wb = load_workbook(excel_path, keep_vba=True)
        ws = wb[EXCEL_SHEET_NAME]
        
        header_row = WEEK_START_ROWS.get(week_num, 7)
        data_start_row = header_row + 1

        # まず7行分クリア（G列=数式は保持）
        for i in range(DATA_ROW_COUNT):
            excel_row = data_start_row + i
            # C-F列クリア
            for col in [COL_WBS, COL_DESCRIPTION, COL_AA_TYPE, 6]:
                ws.cell(row=excel_row, column=col, value=None)
            # H-N列クリア
            for col in [COL_MON, COL_TUE, COL_WED, COL_THU, COL_FRI, COL_SAT, COL_SUN]:
                ws.cell(row=excel_row, column=col, value=None)

        for i, row_data in enumerate(data):
            if i >= DATA_ROW_COUNT:
                break

            excel_row = data_start_row + i

            ws.cell(row=excel_row, column=COL_WBS, value=row_data.get("wbs_element", ""))
            ws.cell(row=excel_row, column=COL_DESCRIPTION, value=row_data.get("description", ""))
            ws.cell(row=excel_row, column=COL_AA_TYPE, value=row_data.get("aa_type", ""))
            
            hours = row_data.get("hours", {})
            ws.cell(row=excel_row, column=COL_MON, value=hours.get("mon", 0) if hours.get("mon", 0) > 0 else None)
            ws.cell(row=excel_row, column=COL_TUE, value=hours.get("tue", 0) if hours.get("tue", 0) > 0 else None)
            ws.cell(row=excel_row, column=COL_WED, value=hours.get("wed", 0) if hours.get("wed", 0) > 0 else None)
            ws.cell(row=excel_row, column=COL_THU, value=hours.get("thu", 0) if hours.get("thu", 0) > 0 else None)
            ws.cell(row=excel_row, column=COL_FRI, value=hours.get("fri", 0) if hours.get("fri", 0) > 0 else None)
            ws.cell(row=excel_row, column=COL_SAT, value=hours.get("sat", 0) if hours.get("sat", 0) > 0 else None)
            ws.cell(row=excel_row, column=COL_SUN, value=hours.get("sun", 0) if hours.get("sun", 0) > 0 else None)
        
        wb.save(excel_path)
        print(f"  ✅ ファイル保存完了: {len(data)}行（再読み込みが必要）")
        return True
        
    except Exception as e:
        _logger.error(f"Excel保存エラー: {e}\n{traceback.format_exc()}")
        print(f"  Excel保存エラー: {e}")
        print("  【対処法】")
        print(f"    1. Excelファイル '{excel_path}' が存在するか確認")
        print("    2. ファイルが他のプログラムでロックされていないか確認")
        return False


def _week_data_signature(data):
    """週データの簡易シグネチャ（週切替失敗検知用）"""
    if not data:
        return "EMPTY"

    rows = []
    for row in data:
        h = row.get("hours", {})
        rows.append(
            (
                row.get("wbs_element", ""),
                row.get("description", ""),
                row.get("aa_type", ""),
                h.get("mon", 0),
                h.get("tue", 0),
                h.get("wed", 0),
                h.get("thu", 0),
                h.get("fri", 0),
                h.get("sat", 0),
                h.get("sun", 0),
            )
        )

    return str(rows)


def load_from_catw_excel(excel_path=None, week_num=None):
    """CATWマクロExcelからデータを読み込み（開いているExcelから優先）"""
    if not week_num:
        week_num = get_current_week_num()
    
    print(f"  Excelから読み込み中: Week {week_num}")
    
    header_row = WEEK_START_ROWS.get(week_num, 7)
    data_start_row = header_row + 1
    
    # win32comで開いているExcelから読み込む（優先）
    if WIN32COM_AVAILABLE:
        try:
            # GetActiveObjectで既存のExcelインスタンスに接続
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
            except Exception as e:
                _logger.debug(f"GetActiveObject失敗(読み込み): {e}")
                excel = win32com.client.Dispatch("Excel.Application")
            
            target_wb = None
            if excel.Workbooks.Count > 0:
                for i in range(1, excel.Workbooks.Count + 1):
                    wb = excel.Workbooks.Item(i)
                    if "CATW" in wb.Name:
                        target_wb = wb
                        break
            
            if target_wb:
                ws = target_wb.Sheets(EXCEL_SHEET_NAME)
                
                data = []
                for i in range(DATA_ROW_COUNT):
                    excel_row = data_start_row + i
                    
                    wbs = ws.Cells(excel_row, COL_WBS).Value or ""
                    aa_type_raw = ws.Cells(excel_row, COL_AA_TYPE).Value or ""
                    
                    # AA Type からコードを抽出 (例: "413 Documentation" → コード="413")
                    aa_type_code = ""
                    aa_type = str(aa_type_raw)
                    if aa_type:
                        parts = aa_type.split(" ", 1)
                        if parts and parts[0].isdigit():
                            aa_type_code = parts[0]
                    
                    hours = {
                        "mon": ws.Cells(excel_row, COL_MON).Value or 0,
                        "tue": ws.Cells(excel_row, COL_TUE).Value or 0,
                        "wed": ws.Cells(excel_row, COL_WED).Value or 0,
                        "thu": ws.Cells(excel_row, COL_THU).Value or 0,
                        "fri": ws.Cells(excel_row, COL_FRI).Value or 0,
                        "sat": ws.Cells(excel_row, COL_SAT).Value or 0,
                        "sun": ws.Cells(excel_row, COL_SUN).Value or 0,
                    }
                    
                    # 数値に変換
                    for day in hours:
                        try:
                            hours[day] = float(hours[day]) if hours[day] else 0.0
                        except (ValueError, TypeError) as e:
                            _logger.debug(f"時間値変換スキップ(win32com行{excel_row} {day}): {e}")
                            hours[day] = 0.0
                    
                    # データがある行のみ追加
                    if wbs or any(v > 0 for v in hours.values()):
                        data.append({
                            "row": i + 1,
                            "wbs_element": str(wbs),
                            "aa_type": aa_type,
                            "aa_type_code": aa_type_code,
                            "hours": hours
                        })
                
                print(f"  読み込み完了（開いているExcelから）: {len(data)}行")
                return data
            else:
                print("  警告: CATWブックが開いていません（win32com経由）。openpyxlで読み込みます...")
                
        except Exception as e:
            _logger.error(f"win32com読み込みエラー: {e}\n{traceback.format_exc()}")
            print(f"  win32com読み込みエラー: {e}")
            print("  【対処法】")
            print("    1. CATWマクロExcelが開いているか確認")
            print(f"    2. シート名 '{EXCEL_SHEET_NAME}' が存在するか確認")
            print("  openpyxlで読み込みを試みます...")
    
    # フォールバック: openpyxlで読み込み
    if not EXCEL_AVAILABLE:
        print("  Excel読み込み機能が利用できません（openpyxlが必要）")
        return []
    
    if not excel_path:
        excel_path = DEFAULT_EXCEL_PATH
    
    if not os.path.exists(excel_path):
        print(f"  Excelファイルが見つかりません: {excel_path}")
        return []
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb[EXCEL_SHEET_NAME]
        
        data = []
        for i in range(DATA_ROW_COUNT):
            excel_row = data_start_row + i
            
            wbs = ws.cell(row=excel_row, column=COL_WBS).value or ""
            aa_type_raw = ws.cell(row=excel_row, column=COL_AA_TYPE).value or ""
            
            # AA Type からコードを抽出 (例: "413 Documentation" → コード="413")
            aa_type_code = ""
            aa_type = str(aa_type_raw)
            if aa_type:
                parts = aa_type.split(" ", 1)
                if parts and parts[0].isdigit():
                    aa_type_code = parts[0]
            
            hours = {
                "mon": ws.cell(row=excel_row, column=COL_MON).value or 0,
                "tue": ws.cell(row=excel_row, column=COL_TUE).value or 0,
                "wed": ws.cell(row=excel_row, column=COL_WED).value or 0,
                "thu": ws.cell(row=excel_row, column=COL_THU).value or 0,
                "fri": ws.cell(row=excel_row, column=COL_FRI).value or 0,
                "sat": ws.cell(row=excel_row, column=COL_SAT).value or 0,
                "sun": ws.cell(row=excel_row, column=COL_SUN).value or 0,
            }
            
            # 数値に変換
            for day in hours:
                try:
                    hours[day] = float(hours[day]) if hours[day] else 0.0
                except (ValueError, TypeError) as e:
                    _logger.debug(f"時間値変換スキップ(openpyxl行{excel_row} {day}): {e}")
                    hours[day] = 0.0
            
            # データがある行のみ追加
            if wbs or any(v > 0 for v in hours.values()):
                data.append({
                    "row": i + 1,
                    "wbs_element": str(wbs),
                    "aa_type": aa_type,
                    "aa_type_code": aa_type_code,
                    "hours": hours
                })
        
        print(f"  読み込み完了（ファイルから）: {len(data)}行")
        return data
        
    except Exception as e:
        _logger.error(f"Excel読み込みエラー: {e}\n{traceback.format_exc()}")
        print(f"  Excel読み込みエラー: {e}")
        print("  【対処法】")
        print(f"    1. Excelファイル '{excel_path}' が存在するか確認")
        print("    2. ファイルが他のプログラムでロックされていないか確認")
        return []


def main():
    """メイン処理"""
    silent_mode = "--silent" in sys.argv
    if silent_mode:
        sys.argv.remove("--silent")
    
    if len(sys.argv) < 2:
        print("使用方法:")
        print("  python catw_playwright.py read [週番号] [Excelパス]   - CATWから読み取り→Excelに保存")
        print("  python catw_playwright.py readmonth [Excelパス]       - 月全体を読み取り")
        print("  python catw_playwright.py input [週番号] [Excelパス]  - Excelから読み取り→CATWに入力")
        print("  python catw_playwright.py inputmonth [Excelパス]      - 月全体を入力")
        print("  python catw_playwright.py connect                     - CATWに接続")
        print("")
        print("オプション:")
        print("  週番号: 1-6（省略時は今日の日付から自動判定）")
        print("  --silent: 対話なしで実行（VBAから呼び出し用）")
        return
    
    command = sys.argv[1]
    
    # 引数の解析
    week_num = None
    excel_path = None
    
    for arg in sys.argv[2:]:
        if arg.isdigit() and 1 <= int(arg) <= 6:
            week_num = int(arg)
        elif arg.endswith(('.xlsx', '.xlsm', '.xls')):
            excel_path = arg
    
    catw = CATWAutomation()
    
    try:
        if command == "connect":
            print("=" * 60)
            print("CATWブラウザを起動します")
            print("=" * 60)
            if not catw.start_debug_browser():
                print("\nエラー: ブラウザの起動・接続に失敗しました")
                print("\n【対処法】")
                print("  1. タスクマネージャーですべてのmsedge.exeを終了してから再実行")
                print("  2. Microsoft Edgeがインストールされているか確認")
                if not silent_mode:
                    input("\n終了するにはEnterキーを押してください...")
                sys.exit(2)
            if not silent_mode:
                print("\nログイン後、左メニューから【Time Entry】をクリックしてから、")
                print("Enterキーを押してください...")
                input()
            print("接続成功！このブラウザは開いたままにしてください。")
        
        elif command == "readmonth":
            print("=" * 60)
            print("CATW データ読み取り（月全体: Week 1-6）")
            print("=" * 60)
            
            if excel_path:
                print(f"保存先Excel: {excel_path}")
            
            if not catw.connect(reuse_browser=True):
                print("\nエラー: ブラウザに接続できません。")
                print("先に「Open CATW」ボタンでブラウザを起動してログインしてください。")
                if not silent_mode:
                    input("\n終了するにはEnterキーを押してください...")
                sys.exit(2)
            
            excel_year, month = get_year_month_from_excel()
            if not month:
                month = extract_month_from_filename(excel_path) if excel_path else None
            
            success_count = 0
            prev_sig = None
            for wk in range(1, 7):
                print(f"\n{'='*40}")
                print(f"Week {wk} 読み取り中...")
                print(f"{'='*40}")
                
                try:
                    if not catw.navigate_to_week(wk, month, year=excel_year):
                        print(f"警告: Week {wk} への移動に失敗しました。")
                        time.sleep(2)
                        continue
                    
                    time.sleep(2)  # ページ読み込み待ち
                    data = catw.read_week_data()

                    # 週切替が効かず同一データを読み続けるケースへの保険
                    current_sig = _week_data_signature(data)
                    if wk > 1 and current_sig == prev_sig and current_sig != "EMPTY":
                        _logger.warning(f"Week {wk}: 前週と同一データを検出。週切替を再試行します。")
                        print("警告: 前週と同一データを検出。週切替を再試行します...")
                        if catw.navigate_to_week(wk, month, year=excel_year):
                            time.sleep(2)
                            data = catw.read_week_data()
                            current_sig = _week_data_signature(data)

                    prev_sig = current_sig
                    
                    if data:
                        print(f"Week {wk}: {len(data)}行のデータを取得")
                        save_to_catw_excel(data, excel_path=excel_path, week_num=wk)
                        success_count += 1
                    else:
                        print(f"Week {wk}: データなし")
                except Exception as e:
                    _logger.error(f"Week {wk} 読み取りエラー: {e}\n{traceback.format_exc()}")
                    print(f"Week {wk} でエラー発生: {e}")
                    continue
            
            print(f"\n{'='*60}")
            print(f"月間データ読み取り完了: {success_count}/6 週成功")
            print(f"{'='*60}")
            
            if not silent_mode:
                input("\n終了するにはEnterキーを押してください...")
        
        elif command == "read":
            if not week_num:
                week_num = get_current_week_num()
            
            print("=" * 60)
            print(f"CATW データ読み取り (Week {week_num})")
            print("=" * 60)
            
            if not catw.connect(reuse_browser=True):
                print("\nエラー: ブラウザに接続できません。")
                if not silent_mode:
                    input("\n終了するにはEnterキーを押してください...")
                sys.exit(2)
            
            excel_year, month = get_year_month_from_excel()
            if not month:
                month = extract_month_from_filename(excel_path) if excel_path else None
            if not catw.navigate_to_week(week_num, month, year=excel_year):
                print("警告: 週の切り替えに失敗しました。")
            
            time.sleep(1)
            data = catw.read_week_data()
            
            if data:
                print(f"\n{len(data)}行のデータを取得しました")
                save_to_catw_excel(data, excel_path=excel_path, week_num=week_num)
            else:
                print("データが見つかりませんでした。")
            
            if not silent_mode:
                input("\n終了するにはEnterキーを押してください...")
        
        elif command == "input":
            # ExcelからCATWにデータを入力
            if not week_num:
                week_num = get_current_week_num()
            
            print("=" * 60)
            print(f"CATW データ入力 (Week {week_num})")
            print("=" * 60)
            
            if excel_path:
                print(f"入力元Excel: {excel_path}")
            
            # Excelからデータを読み込み
            data = load_from_catw_excel(excel_path=excel_path, week_num=week_num)
            if not data:
                print("Excelデータなし（Webの当月分をクリアします）")
                data = []  # 空リストで処理続行
            
            print(f"\n{len(data)}行のデータを入力します")
            
            # ブラウザに接続
            if not catw.connect(reuse_browser=True):
                print("\nエラー: ブラウザに接続できません。")
                if not silent_mode:
                    input("\n終了するにはEnterキーを押してください...")
                sys.exit(2)
            
            excel_year, month = get_year_month_from_excel()
            if not month:
                month = extract_month_from_filename(excel_path) if excel_path else None
            if not catw.navigate_to_week(week_num, month, year=excel_year):
                print("警告: 週の切り替えに失敗しました。")
            
            time.sleep(1)
            
            # データを入力
            if catw.input_week_data(data, week_num=week_num, month=month, year=excel_year):
                print("\nデータ入力完了")
                # 保存
                if catw.save():
                    print("CATWに保存しました")
                else:
                    print("警告: 保存に失敗した可能性があります")
            else:
                print("データ入力に失敗しました")
            
            if not silent_mode:
                input("\n終了するにはEnterキーを押してください...")
        
        elif command == "inputmonth":
            # 月全体をCATWに入力
            print("=" * 60)
            print("CATW データ入力（月全体: Week 1-6）")
            print("=" * 60)
            
            if excel_path:
                print(f"入力元Excel: {excel_path}")
            
            # ブラウザに接続
            if not catw.connect(reuse_browser=True):
                print("\nエラー: ブラウザに接続できません。")
                if not silent_mode:
                    input("\n終了するにはEnterキーを押してください...")
                sys.exit(2)
            
            excel_year, month = get_year_month_from_excel()
            if not month:
                month = extract_month_from_filename(excel_path) if excel_path else None
            
            success_count = 0
            for wk in range(1, 7):
                print(f"\n{'='*40}")
                print(f"Week {wk} 入力中...")
                print(f"{'='*40}")
                
                try:
                    # Excelからデータを読み込み
                    data = load_from_catw_excel(excel_path=excel_path, week_num=wk)
                    if not data:
                        print(f"Week {wk}: Excelデータなし（Webの当月分をクリアします）")
                        data = []  # 空リストで処理続行（既存行の当月列クリア用）
                    
                    # 週に移動
                    if not catw.navigate_to_week(wk, month, year=excel_year):
                        print(f"警告: Week {wk} への移動に失敗しました。")
                        time.sleep(2)
                        continue
                    
                    time.sleep(2)  # ページ読み込み待ち
                    
                    # データを入力
                    if catw.input_week_data(data, week_num=wk, month=month, year=excel_year):
                        if catw.save():
                            print(f"Week {wk}: {len(data)}行入力・保存完了")
                            success_count += 1
                        else:
                            print(f"Week {wk}: 保存に失敗")
                    else:
                        print(f"Week {wk}: 入力に失敗")
                except Exception as e:
                    _logger.error(f"Week {wk} 入力エラー: {e}\n{traceback.format_exc()}")
                    print(f"Week {wk} でエラー発生: {e}")
                    continue
            
            print(f"\n{'='*60}")
            print(f"月間データ入力完了: {success_count}/6 週成功")
            print(f"{'='*60}")
            
            if success_count == 0:
                print("エラー: どの週もデータを入力できませんでした")
                if not silent_mode:
                    input("\n終了するにはEnterキーを押してください...")
                sys.exit(3)
            
            if not silent_mode:
                input("\n終了するにはEnterキーを押してください...")
        
        else:
            print(f"不明なコマンド: {command}")
    
    finally:
        catw.close()


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception as e:
        _logger.error(f"予期しないエラー: {e}\n{traceback.format_exc()}")
        print(f"\n{'='*60}")
        print(f"予期しないエラーが発生しました: {e}")
        print(f"{'='*60}")
        print(f"\n詳細なエラー情報:")
        traceback.print_exc()
        print(f"\n【対処法】")
        print("  1. catw_debug.log ファイルを確認してください")
        print("  2. 上記のエラーメッセージを開発者に報告してください")
        print(f"  3. Python: {sys.executable} ({sys.version})")
        sys.exit(99)