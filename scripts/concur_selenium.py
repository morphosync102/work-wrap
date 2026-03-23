# -*- coding: utf-8 -*-
"""
Concur経費精算システム Selenium自動化スクリプト
Excelから交通費データを読み取り、Concurに自動入力
"""

import sys
import os
import time
from pathlib import Path
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException, StaleElementReferenceException

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: openpyxlがインストールされていません。pip install openpyxl を実行してください。")

# WebDriverのパス
EDGE_DRIVER_PATH = r"C:\Tools\CATW\msedgedriver.exe"

# デバッグポート（既存ブラウザに接続するため）
DEBUG_PORT = 9222

# Excelのシート名
EXCEL_SHEET_NAME = "Concur"



def load_expense_data(excel_path):
    """
    ExcelのConcurシートから経費データを読み込む

    Args:
        excel_path: Excelファイルのパス

    Returns:
        list: 経費データのリスト
        [
            {
                "date": 日付,
                "vendor": 交通手段（B列）,
                "amount": 金額（C列）,
                "comment": コメント（D列）
            },
            ...
        ]
    """
    if not EXCEL_AVAILABLE:
        print("openpyxlが利用できません")
        return []

    try:
        wb = load_workbook(excel_path, data_only=True)

        if EXCEL_SHEET_NAME not in wb.sheetnames:
            print(f"エラー: シート '{EXCEL_SHEET_NAME}' が見つかりません")
            print(f"利用可能なシート: {wb.sheetnames}")
            return []

        ws = wb[EXCEL_SHEET_NAME]
        expenses = []

        # A2から開始（A1はヘッダーと想定）
        row = 2
        while True:
            date_val = ws.cell(row=row, column=1).value  # A列: 日付
            vendor = ws.cell(row=row, column=2).value     # B列: 交通手段
            amount = ws.cell(row=row, column=3).value     # C列: 金額
            comment = ws.cell(row=row, column=4).value    # D列: コメント
            purpose = ws.cell(row=row, column=5).value    # E列: Business Purpose

            # A列が空ならループ終了
            if date_val is None or str(date_val).strip() == "":
                break

            # 日付のフォーマット変換
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%m/%d/%Y")
            else:
                date_str = str(date_val)

            expenses.append({
                "date": date_str,
                "vendor": str(vendor) if vendor else "",
                "amount": str(amount) if amount else "0",
                "comment": str(comment) if comment else "",
                "purpose": str(purpose) if purpose else ""
            })

            row += 1

        wb.close()
        print(f"読み込み完了: {len(expenses)}件の経費データ")
        return expenses

    except Exception as e:
        print(f"Excelの読み込みに失敗: {e}")
        return []



class ConcurAutomation:
    """Concur経費精算システム自動化クラス"""


    def __init__(self):
        self.driver = None


    def connect(self):
        """既存のEdgeブラウザに接続"""
        options = Options()
        options.add_experimental_option("debuggerAddress", f"127.0.0.1:{DEBUG_PORT}")

        if EDGE_DRIVER_PATH and os.path.exists(EDGE_DRIVER_PATH):
            service = Service(executable_path=EDGE_DRIVER_PATH)
        else:
            service = Service()

        try:
            self.driver = webdriver.Edge(service=service, options=options)
            print(f"ブラウザに接続しました")
            print(f"現在のページ: {self.driver.title}")
            return True
        except Exception as e:
            print(f"ブラウザ接続エラー: {e}")
            print("ヒント: 先にEdgeをデバッグモードで起動してください")
            print(f'  msedge --remote-debugging-port={DEBUG_PORT} "ConcurのURL"')
            return False


    def wait_and_click(self, xpath, timeout=10, description=""):
        """要素を待機してクリック"""
        def _click_element(el):
            click_target = el
            try:
                tag = (el.tag_name or "").lower()
            except Exception:
                tag = ""

            if tag not in ("button", "a"):
                try:
                    ancestors = el.find_elements(
                        By.XPATH,
                        "./ancestor::button[1] | ./ancestor::a[1] | ./ancestor::*[@role='button'][1]",
                    )
                    if ancestors:
                        click_target = ancestors[0]
                except Exception:
                    pass

            try:
                click_target.click()
            except ElementClickInterceptedException:
                self.driver.execute_script("arguments[0].click();", click_target)
            except StaleElementReferenceException:
                raise

        def _try_click_in_current_context():
            element = WebDriverWait(self.driver, timeout).until(
                EC.element_to_be_clickable((By.XPATH, xpath))
            )
            _click_element(element)
            return True

        last_error = None

        # まずは現在のフレーム（通常はdefault content）で試す
        try:
            if _try_click_in_current_context():
                if description:
                    print(f"  ✓ {description}")
                return True
        except Exception as e:
            last_error = e

        # iframe内に要素があるケースを吸収（Concurでよくある）
        try:
            self.driver.switch_to.default_content()
            frames = self.driver.find_elements(By.TAG_NAME, "iframe")
            for frame in frames:
                try:
                    self.driver.switch_to.default_content()
                    self.driver.switch_to.frame(frame)
                    if _try_click_in_current_context():
                        self.driver.switch_to.default_content()
                        if description:
                            print(f"  ✓ {description}")
                        return True
                except Exception as e:
                    last_error = e
                    continue
        finally:
            try:
                self.driver.switch_to.default_content()
            except Exception:
                pass

        print(f"  ✗ クリック失敗 ({description}): {last_error}")
        return False


    def wait_and_input(self, xpath, text, timeout=10, description="", clear=True):
        """要素を待機して入力"""
        try:
            element = WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            if clear:
                element.clear()
            element.send_keys(text)
            if description:
                print(f"  ✓ {description}: {text}")
            return True
        except Exception as e:
            print(f"  ✗ 入力失敗 ({description}): {e}")
            return False


    def select_dropdown_option(self, dropdown_xpath, option_text, timeout=10, description=""):
        """ドロップダウンを開いて選択肢をクリック"""
        try:
            # ドロップダウンをクリック
            dropdown = WebDriverWait(self.driver, timeout).until(
                EC.element_to_be_clickable((By.XPATH, dropdown_xpath))
            )
            dropdown.click()
            time.sleep(0.5)

            # 選択肢をクリック（テキストを含む要素を探す）
            option_xpath = f"//*[contains(text(), '{option_text}')]"
            option = WebDriverWait(self.driver, timeout).until(
                EC.element_to_be_clickable((By.XPATH, option_xpath))
            )
            option.click()

            if description:
                print(f"  ✓ {description}: {option_text}")
            return True
        except Exception as e:
            print(f"  ✗ 選択失敗 ({description}): {e}")
            return False


    def create_new_report(self):
        """
        Step 1-4: 新しいレポートを作成
        """
        print("\n" + "=" * 50)
        print("レポート作成を開始します")
        print("=" * 50)

        # 1. Create Expense Report をクリック（旧UI: Create New Report）
        if not self.wait_and_click(
            "//button[.//*[normalize-space(.)='Create Expense Report'] or contains(normalize-space(.), 'Create Expense Report') or @aria-label='Create Expense Report']"
            " | //*[@role='button' and (contains(normalize-space(.), 'Create Expense Report') or @aria-label='Create Expense Report')]"
            " | //a[contains(normalize-space(.), 'Create Expense Report') or @aria-label='Create Expense Report']"
            " | //button[.//*[normalize-space(.)='Create New Report'] or contains(normalize-space(.), 'Create New Report') or @aria-label='Create New Report']"
            " | //*[@role='button' and (contains(normalize-space(.), 'Create New Report') or @aria-label='Create New Report')]"
            " | //a[contains(normalize-space(.), 'Create New Report') or @aria-label='Create New Report']"
            " | //span[contains(@class, 'sapcnqr-button__text') and (normalize-space(.)='Create Expense Report' or normalize-space(.)='Create New Report')]",
            timeout=20,
            description="Create Expense Report",
        ):
            return False

        time.sleep(2)  # ダイアログ表示待ち

        # 2. Report Name を入力
        current_month = datetime.now().month
        report_name = f"{current_month}月分の交通費精算"

        # Report Nameの入力欄を探す（ラベルの近くのinput、またはname/id属性など）
        report_name_xpaths = [
            "//label[contains(text(), 'Report Name')]/following::input[1]",
            "//input[contains(@placeholder, 'Report Name')]",
            "//input[contains(@name, 'reportName')]",
            "//input[contains(@id, 'reportName')]"
        ]

        input_success = False
        for xpath in report_name_xpaths:
            if self.wait_and_input(xpath, report_name, timeout=3, description="Report Name"):
                input_success = True
                break

        if not input_success:
            print("  警告: Report Name 入力欄が見つかりませんでした")

        # 3. Business Purpose で "Non Travel Expenses" を選択
        # 検索ボックス付きのドロップダウン: クリック → 検索文字入力 → 選択
        try:
            # ドロップダウンをクリックして開く
            purpose_dropdown = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH,
                    "//label[contains(text(), 'Business Purpose')]/following::*[contains(@class, 'select') or contains(@class, 'dropdown') or @role='combobox'][1] | //*[contains(@id, 'businessPurpose')] | //*[@aria-label='Business Purpose']"
                ))
            )
            purpose_dropdown.click()
            time.sleep(0.5)

            # 検索ボックスに入力
            search_box = WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.XPATH,
                    "//input[contains(@placeholder, 'Search')] | //input[contains(@class, 'search')] | //input[@type='search']"
                ))
            )
            search_box.clear()
            search_box.send_keys("Non Travel")
            time.sleep(1)

            # 選択肢をクリック
            option = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.XPATH,
                    "//*[contains(text(), 'Non Travel Expenses')]"
                ))
            )
            option.click()
            print("  ✓ Business Purpose: Non Travel Expenses")
        except Exception as e:
            print(f"  ✗ Business Purpose選択失敗: {e}")

        time.sleep(1)

        # 4. Create Report をクリック
        if not self.wait_and_click(
            "//span[@class='sapcnqr-button__text' and text()='Create Report'] | //button[normalize-space()='Create Report'] | //button//span[text()='Create Report']",
            description="Create Report"
        ):
            return False

        time.sleep(3)  # ページ遷移待ち
        print("レポート作成完了")
        return True


    def add_expense(self, expense_data):
        """
        Step 5-8: 1件の経費を追加

        Args:
            expense_data: {
                "date": 日付,
                "vendor": 交通手段,
                "amount": 金額,
                "comment": コメント
            }
        """
        print(f"\n--- 経費追加: {expense_data['date']} / {expense_data['vendor']} / \{expense_data['amount']} ---")

        # 5. Add Expense → Manually Create Expense
        if not self.wait_and_click(
            "//button[contains(text(), 'Add Expense')] | //*[contains(text(), 'Add Expense')]",
            description="Add Expense"
        ):
            return False

        time.sleep(2)  # 待機時間を増やす

        if not self.wait_and_click(
            "//*[contains(text(), 'Manually Create')] | //*[contains(text(), 'Create New Expense')]",
            description="Manually Create Expense"
        ):
            return False

        time.sleep(3)  # 待機時間を増やす

        # 6. Public Transport を選択
        # ユーザー提供: <span class="expense-type-list__expense-type-button-text sapcnqr-button__text">Public Transport</span>
        if not self.wait_and_click(
            "//span[contains(@class, 'expense-type-list__expense-type-button-text') and contains(text(), 'Public Transport')] | //span[@class='sapcnqr-button__text' and contains(text(), 'Public Transport')]",
            timeout=10,
            description="Public Transport"
        ):
            return False

        time.sleep(2)

        # 7. 各項目を入力

        # Transaction Date
        date_xpaths = [
            "//label[contains(text(), 'Transaction Date')]/following::input[1]",
            "//input[contains(@name, 'transactionDate')]",
            "//input[contains(@id, 'date')]"
        ]
        for xpath in date_xpaths:
            if self.wait_and_input(xpath, expense_data["date"], timeout=3, description="Transaction Date"):
                break

        # Transport Type: Searchable Dropdown (Bus, Ferry, Subway, etc.)
        # Default to Subway
        transport_choice = "Subway"
        if expense_data.get("vendor"):
            v_lower = str(expense_data["vendor"]).lower()
            if "bus" in v_lower or "バス" in v_lower:
                transport_choice = "Bus"
            elif "ferry" in v_lower or "フェリー" in v_lower:
                transport_choice = "Ferry"
            elif "taxi" in v_lower or "タクシー" in v_lower:
                transport_choice = "Taxi" # If Taxi exists in list, otherwise might be Subway

        try:
            # Dropdown click
            t_drop = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.XPATH,
                    "//label[contains(text(), 'Transport Type')]/following::*[contains(@class, 'select') or contains(@class, 'dropdown') or @role='combobox'][1] | //*[contains(@id, 'transportType')]"
                ))
            )
            t_drop.click()
            time.sleep(0.5)

            # Search Input
            t_search = WebDriverWait(self.driver, 3).until(
                EC.presence_of_element_located((By.XPATH,
                    "//input[contains(@placeholder, 'Search')]"
                ))
            )
            t_search.clear()
            t_search.send_keys(transport_choice)
            time.sleep(0.5)

            # Option click
            t_opt = WebDriverWait(self.driver, 3).until(
                EC.element_to_be_clickable((By.XPATH,
                    f"//*[contains(text(), '{transport_choice}')]"
                ))
            )
            t_opt.click()
            print(f"  ✓ Transport Type: {transport_choice}")
        except Exception as e:
            print(f"  ✗ Transport Type選択失敗: {e}")

        # Enter Vendor Name
        vendor_xpaths = [
            "//label[contains(text(), 'Vendor')]/following::input[1]",
            "//input[contains(@placeholder, 'Vendor')]",
            "//input[contains(@name, 'vendor')]"
        ]
        for xpath in vendor_xpaths:
            if self.wait_and_input(xpath, expense_data["vendor"], timeout=3, description="Vendor Name"):
                break

        # City of Purchase: "Tokyo, Tokyo"
        # ユーザー提供HTML:
        # 枠: <span id="location" class="sapcnqr-select-field__input">
        # 入力欄: <input data-nuiexp="field-location__input">
        city_target = "Tokyo, Tokyo"
        try:
            # 1. 枠（id="location"）をクリック
            city_field = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.XPATH,
                    "//*[@id='location']"
                ))
            )
            city_field.click()
            time.sleep(1.0)

            # 2. 入力欄（data-nuiexp="field-location__input"）に入力
            c_input = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.XPATH,
                    "//input[@data-nuiexp='field-location__input']"
                ))
            )
            c_input.click()
            c_input.clear()
            c_input.send_keys("Tokyo")
            time.sleep(1.5)

            # 3. 候補リストから直接クリック
            # ユーザー提供: <div class="sapcnqr-listbox-item__wrapper">Tokyo, Tokyo</div>
            c_opt = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.XPATH,
                    "//div[@class='sapcnqr-listbox-item__wrapper' and contains(text(), 'Tokyo')]"
                ))
            )
            c_opt.click()
            time.sleep(0.5)

            print(f"  ✓ City of Purchase: {city_target}")
        except Exception as e:
            print(f"  ✗ City選択失敗: {e}")

        # Amount
        amount_xpaths = [
            "//label[contains(text(), 'Amount')]/following::input[1]",
            "//input[contains(@name, 'amount')]",
            "//input[contains(@id, 'amount')]"
        ]
        for xpath in amount_xpaths:
            if self.wait_and_input(xpath, expense_data["amount"], timeout=3, description="Amount"):
                break

        # Business Purpose (Excel E列)
        if expense_data.get("purpose"):
            purpose_text = expense_data["purpose"]
            try:
                # Dropdown click
                p_drop = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH,
                        "//label[contains(text(), 'Business Purpose')]/following::*[contains(@class, 'select') or contains(@class, 'dropdown') or @role='combobox'][1]"
                    ))
                )
                p_drop.click()
                time.sleep(0.5)

                # Search Input
                p_search = WebDriverWait(self.driver, 3).until(
                    EC.presence_of_element_located((By.XPATH,
                        "//input[contains(@placeholder, 'Search')]"
                    ))
                )
                p_search.clear()
                p_search.send_keys(purpose_text)
                time.sleep(1.0)

                # Option click (Transport Typeと同様のロジック)
                p_opt = WebDriverWait(self.driver, 3).until(
                    EC.element_to_be_clickable((By.XPATH,
                        f"//*[contains(text(), '{purpose_text}')]"
                    ))
                )
                p_opt.click()
                print(f"  ✓ Business Purpose: {purpose_text}")
            except Exception as e:
                print(f"  ✗ Business Purpose選択失敗: {e}")

        time.sleep(1)

        # Receipt Status: Tax Receipt
        receipt_xpaths = [
            "//label[contains(text(), 'Receipt')]/following::*[contains(@class, 'select')][1]",
            "//*[contains(@id, 'receiptStatus')]"
        ]
        for xpath in receipt_xpaths:
            if self.select_dropdown_option(xpath, "Tax Receipt", timeout=3, description="Receipt Status"):
                break

        # Comment
        comment_xpaths = [
            "//label[contains(text(), 'Comment')]/following::textarea[1]",
            "//label[contains(text(), 'Comment')]/following::input[1]",
            "//textarea[contains(@name, 'comment')]"
        ]
        for xpath in comment_xpaths:
            if self.wait_and_input(xpath, expense_data["comment"], timeout=3, description="Comment"):
                break

        time.sleep(1)

        # 8. Save Expense
        if not self.wait_and_click(
            "//span[@class='sapcnqr-button__text' and contains(text(), 'Save Expense')] | //button[normalize-space()='Save Expense'] | //button//span[contains(text(), 'Save Expense')]",
            description="Save Expense"
        ):
            return False

        time.sleep(3)  # 保存完了待ち
        print("  経費を保存しました")
        return True


    def process_all_expenses(self, excel_path):
        """
        全経費を処理

        Args:
            excel_path: Excelファイルのパス
        """
        # Excelからデータ読み込み
        expenses = load_expense_data(excel_path)

        if not expenses:
            print("処理する経費データがありません")
            return False

        print(f"\n{len(expenses)}件の経費を入力します")

        # 新規レポート作成
        if not self.create_new_report():
            print("レポート作成に失敗しました")
            return False

        # 各経費を追加
        success_count = 0
        for i, expense in enumerate(expenses, 1):
            print(f"\n[{i}/{len(expenses)}]")
            if self.add_expense(expense):
                success_count += 1
            else:
                print(f"  警告: 経費 {i} の追加に失敗しました")

        print("\n" + "=" * 50)
        print(f"処理完了: {success_count}/{len(expenses)}件 成功")
        print("=" * 50)
        print("\n★ 最終確認後、手動で Submit してください ★")

        return True



def main():
    """メイン関数"""
    print("=" * 60)
    print("Concur 交通費精算 自動入力ツール")
    print("=" * 60)

    # 引数チェック
    if len(sys.argv) < 2:
        print("\n使用方法:")
        print("  python concur_selenium.py <Excelファイルパス>")
        print("\n例:")
        print('  python concur_selenium.py "C:\\Users\\ishigaki\\Desktop\\経費精算.xlsx"')
        print("\n注意:")
        print("  - 事前にEdgeをデバッグモードで起動し、Concurにログインしてください")
        print(f'  - msedge --remote-debugging-port={DEBUG_PORT} "ConcurのURL"')
        print("  - ExcelにはConcurシートを作成し、A2以降にデータを入力してください")
        print("    A列: 日付, B列: 交通手段, C列: 金額, D列: コメント")
        return

    excel_path = sys.argv[1]

    if not os.path.exists(excel_path):
        print(f"エラー: ファイルが見つかりません: {excel_path}")
        return

    # 自動化開始
    concur = ConcurAutomation()

    if not concur.connect():
        print("\nブラウザに接続できませんでした。")
        print("先にEdgeをデバッグモードで起動してConcurにログインしてください。")
        input("\n終了するにはEnterキーを押してください...")
        return

    # 確認
    print(f"\n使用Excel: {excel_path}")
    confirm = input("処理を開始しますか？ (y/n): ")
    if confirm.lower() != 'y':
        print("キャンセルしました")
        return

    # 処理実行
    concur.process_all_expenses(excel_path)

    input("\n終了するにはEnterキーを押してください...")

if __name__ == "__main__":
    main()
