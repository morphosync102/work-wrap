#!/usr/bin/env python3
"""
MD → Concur Excel 転記スクリプト

MDファイルの「## Concur（交通費）」テーブルを読み込み、
config.yaml で指定した Concur Excel シートに転記します。

使い方:
  python scripts/transfer_concur.py 2026 3   # 特定の年月
  python scripts/transfer_concur.py           # 引数なしで今月

Excel 構造（Concur シート）:
  A=日付 / B=交通手段 / C=金額 / D=コメント / E=Business Purpose / F=入力ステータス
  1行目: ヘッダー / 2行目〜: データ
"""

import sys
import re
from datetime import date
from pathlib import Path

import yaml
from openpyxl import load_workbook


def load_config(path: Path = Path("config.yaml")) -> dict:
    if not path.exists():
        print(f"[エラー] 設定ファイルが見つかりません: {path}")
        print("  cp config.example.yaml config.yaml  を実行して設定してください。")
        sys.exit(1)
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def parse_concur_table(md_text: str, year: int, month: int) -> list[dict]:
    """MDの Concur テーブルをパースする。

    Returns:
        [{date, transport, amount, comment, purpose}]
    """
    m = re.search(r"## Concur（交通費）\n(.*?)(?=\n## |\Z)", md_text, re.DOTALL)
    if not m:
        return []

    section = m.group(1)
    entries = []

    for line in section.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if not cells or cells[0] in ("日付", "------", ""):
            continue
        if re.match(r"-{3,}", cells[0]):
            continue
        # 全セル空の行はスキップ
        if all(c == "" for c in cells):
            continue

        date_str = cells[0]
        transport = cells[1] if len(cells) > 1 else ""
        amount_str = cells[2] if len(cells) > 2 else ""
        comment = cells[3] if len(cells) > 3 else ""
        purpose = cells[4] if len(cells) > 4 else ""

        # 日付解析: YYYY-MM-DD / YYYY/MM/DD / M/D
        dm = re.match(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", date_str)
        if dm:
            try:
                d = date(int(dm.group(1)), int(dm.group(2)), int(dm.group(3)))
            except ValueError:
                continue
        else:
            dm2 = re.match(r"(\d{1,2})/(\d{1,2})", date_str)
            if dm2:
                try:
                    d = date(year, int(dm2.group(1)), int(dm2.group(2)))
                except ValueError:
                    continue
            else:
                continue

        # 金額解析（カンマ・¥記号を除去）
        amount = 0
        amount_clean = re.sub(r"[,¥￥\s]", "", amount_str)
        try:
            amount = int(amount_clean)
        except ValueError:
            pass

        if not transport and amount == 0:
            continue

        entries.append({
            "date": d,
            "transport": transport,
            "amount": amount,
            "comment": comment,
            "purpose": purpose,
        })

    return entries


def write_concur_excel(excel_path: str, sheet_name: str, entries: list, data_start_row: int) -> None:
    """entries を Concur Excel シートに書き込む。"""
    wb = load_workbook(excel_path, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        print(f"[エラー] シート '{sheet_name}' が見つかりません（存在するシート: {wb.sheetnames}）")
        sys.exit(1)

    ws = wb[sheet_name]

    for i, entry in enumerate(entries):
        r = data_start_row + i
        ws.cell(row=r, column=1, value=entry["date"])
        ws.cell(row=r, column=2, value=entry["transport"])
        ws.cell(row=r, column=3, value=entry["amount"])
        ws.cell(row=r, column=4, value=entry["comment"])
        ws.cell(row=r, column=5, value=entry["purpose"])

    wb.save(excel_path)
    print(f"✅ Concur Excel に転記しました: {excel_path}（{len(entries)} 件）")


def main() -> None:
    today = date.today()
    if len(sys.argv) == 3:
        year, month = int(sys.argv[1]), int(sys.argv[2])
    elif len(sys.argv) == 1:
        year, month = today.year, today.month
    else:
        print("使い方: python scripts/transfer_concur.py [年] [月]")
        sys.exit(1)

    config = load_config()

    md_path = Path("勤怠") / str(year) / f"{month}月.md"
    if not md_path.exists():
        print(f"[エラー] MDファイルが見つかりません: {md_path}")
        sys.exit(1)

    md_text = md_path.read_text(encoding="utf-8")
    entries = parse_concur_table(md_text, year, month)

    if not entries:
        print("[情報] Concur テーブルにデータがありません。転記をスキップします。")
        sys.exit(0)

    concur_cfg = config.get("concur", {}).get("excel", {})
    excel_path = concur_cfg.get("path", "")
    sheet_name = concur_cfg.get("sheet_name", "Concur")
    data_start_row = int(concur_cfg.get("data_start_row", 2))

    if not excel_path or "/path/to/" in excel_path:
        print("[エラー] config.yaml の concur.excel.path を実際のExcelファイルパスに設定してください")
        sys.exit(1)

    if not Path(excel_path).exists():
        print(f"[エラー] Excelファイルが見つかりません: {excel_path}")
        sys.exit(1)

    print(f"\n転記対象: {year}年{month}月 → {excel_path}（{sheet_name} シート）")
    for e in entries:
        print(f"  {e['date']}  {e['transport']:<30}  ¥{e['amount']:>6,}  {e['comment']}")
    total_amount = sum(e["amount"] for e in entries)
    print(f"合計: {len(entries)} 件 / ¥{total_amount:,}")

    ans = input("\n転記を実行しますか？ [y/N]: ").strip().lower()
    if ans != "y":
        print("キャンセルしました。")
        sys.exit(0)

    write_concur_excel(excel_path, sheet_name, entries, data_start_row)


if __name__ == "__main__":
    main()
